#!/usr/bin/env python3
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from memo_common import avg, get_num, pct_change, safe_div


REQUIRED_PHASE1_SHEETS = {
    "Summary",
    "Assumptions",
    "IS",
    "BS",
    "CF",
    "Returns",
    "Cross_Check",
    "Raw_Info",
    "_Registry",
    "_State",
}

KNOWN_PHASE2_FILES = {
    "valuation_summary": ["valuation_summary.json", "target_price_summary.json", "target_price.json"],
    "valuation_prep": ["valuation_prep.json"],
    "capital_cost": ["capital_cost.json"],
    "dcf": ["dcf_output.json", "dcf.json"],
    "comps": ["comps_output.json", "comps.json"],
    "reverse_dcf": ["reverse_dcf.json"],
    "football_field": ["football_field.json"],
    "valuation_qc": ["valuation_qc.json"],
    "financials_prep": ["financials_prep.json"],
    "pb_roe": ["pb_roe_output.json"],
    "residual_income": ["residual_income_output.json"],
    "embedded_value": ["embedded_value_output.json"],
    "financials_qc": ["financials_qc.json"],
    "property_bridge": ["property_bridge.json"],
    "reit_nav": ["nav_output.json"],
    "affo_output": ["affo_output.json"],
    "reit_qc": ["reit_qc.json"],
    "regulatory_bridge": ["regulatory_bridge.json"],
    "rab_output": ["rab_output.json"],
    "ddm_output": ["ddm_output.json"],
    "regulated_qc": ["regulated_qc.json"],
    "reserve_model": ["reserve_model.json"],
    "asset_nav_output": ["asset_nav_output.json"],
    "commodity_sensitivity": ["commodity_sensitivity.json"],
    "pnav_market_check": ["pnav_market_check.json"],
    "asset_nav_qc": ["asset_nav_qc.json"],
    "pipeline_registry": ["pipeline_registry.json"],
    "pipeline_rnpv_output": ["pipeline_rnpv_output.json"],
    "cash_runway_dilution": ["cash_runway_dilution.json"],
    "launch_scenarios": ["launch_scenarios.json"],
    "biotech_qc": ["biotech_qc.json"],
    "segment_normalizer": ["segment_normalizer.json"],
    "segment_method_router": ["segment_method_router.json"],
    "sotp_output": ["sotp_output.json"],
    "holdco_discount": ["holdco_discount.json"],
    "sotp_qc": ["sotp_qc.json"],
    "peer_set": ["_peer_set.json"],
    "valuation_state": ["_valuation_state.json"],
    "valuation_log": ["_valuation_log.md"],
}


def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def is_phase1_workbook(path):
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        return REQUIRED_PHASE1_SHEETS.issubset(set(wb.sheetnames))
    except Exception:
        return False


def discover_phase1_workbook(root):
    root = Path(root)
    candidates = []
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if is_phase1_workbook(path):
            candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.stat().st_mtime)


def discover_phase2_files(root):
    root = Path(root)
    found = {}
    for key, names in KNOWN_PHASE2_FILES.items():
        candidates = []
        for name in names:
            candidates.extend(root.rglob(name))
        if candidates:
            found[key] = str(max(candidates, key=lambda item: item.stat().st_mtime))
    return found


def deep_update(base, overlay):
    for key, value in overlay.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_update(base[key], value)
        else:
            base[key] = value
    return base


class WorkbookEvaluator:
    def __init__(self, workbook_path):
        self.workbook_path = str(workbook_path)
        self.wb = load_workbook(workbook_path, data_only=False, read_only=False)
        self.cache = {}
        self.visiting = set()

    def cell_value(self, sheet_name, cell_ref):
        key = (sheet_name, cell_ref)
        if key in self.cache:
            return self.cache[key]
        if key in self.visiting:
            raise ValueError(f"Circular reference detected at {sheet_name}!{cell_ref}")
        self.visiting.add(key)
        ws = self.wb[sheet_name]
        cell = ws[cell_ref]
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            result = self.eval_formula(sheet_name, value)
        else:
            result = value
        self.visiting.remove(key)
        self.cache[key] = result
        return result

    def range_values(self, sheet_name, start_ref, end_ref):
        start_col, start_row = split_cell_ref(start_ref)
        end_col, end_row = split_cell_ref(end_ref)
        values = []
        for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
            for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
                values.append(self.cell_value(sheet_name, f"{get_column_letter(col)}{row}"))
        return values

    def eval_formula(self, current_sheet, formula):
        expression = formula[1:].strip()
        if expression.upper().startswith("IFERROR(") and expression.endswith(")"):
            inner = expression[len("IFERROR("):-1]
            left, right = split_top_level_args(inner)
            try:
                return self.eval_expression(current_sheet, left)
            except Exception:
                return self.eval_expression(current_sheet, right)
        return self.eval_expression(current_sheet, expression)

    def eval_expression(self, current_sheet, expression):
        placeholder_map = {}

        def replace_ranges(text):
            pattern = re.compile(
                r"(?P<sheet1>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ ]*)!)?"
                r"(?P<start>\$?[A-Z]+\$?\d+):"
                r"(?P<sheet2>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ ]*)!)?"
                r"(?P<end>\$?[A-Z]+\$?\d+)"
            )

            def repl(match):
                sheet1 = clean_sheet_name(match.group("sheet1")) or current_sheet
                sheet2 = clean_sheet_name(match.group("sheet2")) or sheet1
                token = f"__RANGE_{len(placeholder_map)}__"
                placeholder_map[token] = f'RANGE("{sheet1}", "{normalize_ref(match.group("start"))}", "{sheet2}", "{normalize_ref(match.group("end"))}")'
                return token

            return pattern.sub(repl, text)

        def replace_cells(text):
            pattern = re.compile(
                r"(?<![A-Za-z0-9_])"
                r"(?P<sheet>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ ]*)!)?"
                r"(?P<cell>\$?[A-Z]+\$?\d+)"
            )

            def repl(match):
                token = f"__CELL_{len(placeholder_map)}__"
                sheet = clean_sheet_name(match.group("sheet")) or current_sheet
                placeholder_map[token] = f'CELL("{sheet}", "{normalize_ref(match.group("cell"))}")'
                return token

            return pattern.sub(repl, text)

        expr = expression.replace("^", "**")
        expr = replace_ranges(expr)
        expr = replace_cells(expr)
        for token, replacement in placeholder_map.items():
            expr = expr.replace(token, replacement)

        env = {
            "CELL": lambda sheet, ref: zero_default(self.cell_value(sheet, ref)),
            "RANGE": lambda sheet1, start_ref, sheet2, end_ref: [
                zero_default(value) for value in self.range_values(sheet1, start_ref, end_ref)
            ],
            "SUM": lambda *args: sum(flatten(args)),
            "AVERAGE": lambda *args: average(flatten(args)),
            "MAX": lambda *args: max(flatten(args)),
            "MIN": lambda *args: min(flatten(args)),
            "ABS": lambda value: abs(zero_default(value)),
            "ROUND": round,
            "IF": lambda cond, true_value, false_value: true_value if cond else false_value,
            "TRUE": True,
            "FALSE": False,
        }
        return eval(expr, {"__builtins__": {}}, env)


def clean_sheet_name(value):
    if not value:
        return None
    return value[:-1].strip("'")


def normalize_ref(value):
    return value.replace("$", "")


def split_cell_ref(ref):
    ref = normalize_ref(ref)
    match = re.match(r"([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(f"Invalid cell ref: {ref}")
    return column_index_from_string(match.group(1)), int(match.group(2))


def split_top_level_args(text):
    level = 0
    parts = []
    last_index = 0
    for index, char in enumerate(text):
        if char == "(":
            level += 1
        elif char == ")":
            level -= 1
        elif char == "," and level == 0:
            parts.append(text[last_index:index].strip())
            last_index = index + 1
    parts.append(text[last_index:].strip())
    if len(parts) != 2:
        raise ValueError(f"Could not split IFERROR arguments: {text}")
    return parts[0], parts[1]


def flatten(values):
    output = []
    for value in values:
        if isinstance(value, list):
            output.extend(flatten(value))
        else:
            output.append(zero_default(value))
    return output


def average(values):
    flat = flatten(values)
    return sum(flat) / len(flat) if flat else 0.0


def zero_default(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    return float(value)


def infer_sheet_periods(ws, header_row, start_col=2):
    periods = []
    for col in range(start_col, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value in (None, ""):
            break
        periods.append((col, str(value).strip()))
    return periods


def read_sheet_rows(evaluator, sheet_name, header_row, start_row, start_col=2):
    ws = evaluator.wb[sheet_name]
    periods = infer_sheet_periods(ws, header_row, start_col=start_col)
    rows = {}
    for row in range(start_row, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label in (None, ""):
            continue
        label = str(label).strip()
        values = {}
        for col, period in periods:
            values[period] = evaluator.cell_value(sheet_name, f"{get_column_letter(col)}{row}")
        rows[label] = values
    return [period for _, period in periods], rows


def actual_periods(periods):
    result = [period for period in periods if period.upper().endswith("A")]
    return result or periods


def forecast_periods(periods):
    actual = set(actual_periods(periods))
    return [period for period in periods if period not in actual]


def last_value(row_map, label, periods):
    row = row_map.get(label, {})
    for period in reversed(periods):
        value = row.get(period)
        if value is not None:
            return float(value)
    return None


def series_values(row_map, label, periods):
    row = row_map.get(label, {})
    values = []
    for period in periods:
        value = row.get(period)
        if value is not None:
            values.append(float(value))
    return values


def parse_title_for_company(summary_title, fallback_name):
    if not summary_title:
        return fallback_name, None
    text = str(summary_title)
    ticker_match = re.search(r"\((?:[A-Z]+:\s*)?([A-Z0-9.\-]+)\)", text)
    ticker = ticker_match.group(1) if ticker_match else None
    name = re.sub(r"\s*\((?:[A-Z]+:\s*)?[A-Z0-9.\-]+\)", "", text)
    name = re.sub(r"\s+[—-]\s+.*$", "", name).strip()
    return name or fallback_name, ticker


def extract_percentages(text):
    return [float(match) / 100.0 for match in re.findall(r"(\d+(?:\.\d+)?)%", str(text))]


def build_default_monitoring(model_summary):
    revenue_base = model_summary.get("revenue_growth_next_year") or model_summary.get("revenue_cagr_forecast")
    ebit_margin_base = model_summary.get("ebit_margin_terminal") or model_summary.get("ebit_margin_ltm")
    roic_base = model_summary.get("roic_terminal") or model_summary.get("roic_ltm")
    drivers = []
    if revenue_base is not None:
        drivers.append(
            {
                "name": "Revenue growth",
                "metric": "revenue_growth",
                "current": model_summary.get("revenue_growth_ltm"),
                "base": revenue_base,
                "bull": revenue_base * 1.15,
                "bear": max(0.0, revenue_base * 0.75),
                "direction": "higher_better",
            }
        )
    if ebit_margin_base is not None:
        drivers.append(
            {
                "name": "EBIT margin",
                "metric": "ebit_margin",
                "current": model_summary.get("ebit_margin_ltm"),
                "base": ebit_margin_base,
                "bull": min(1.0, ebit_margin_base + 0.02),
                "bear": max(0.0, ebit_margin_base - 0.03),
                "direction": "higher_better",
            }
        )
    if roic_base is not None:
        drivers.append(
            {
                "name": "ROIC",
                "metric": "roic",
                "current": model_summary.get("roic_ltm"),
                "base": roic_base,
                "bull": min(1.0, roic_base + 0.02),
                "bear": max(0.0, roic_base - 0.03),
                "direction": "higher_better",
            }
        )

    risks = []
    if model_summary.get("net_debt_to_ebitda") is not None:
        risks.append(
            {
                "name": "Net debt / EBITDA",
                "metric": "net_debt_to_ebitda",
                "current": model_summary["net_debt_to_ebitda"],
                "warning": 3.0,
                "breach": 4.0,
                "direction": "lower_better",
            }
        )
    if model_summary.get("interest_coverage") is not None:
        risks.append(
            {
                "name": "Interest coverage",
                "metric": "interest_coverage",
                "current": model_summary["interest_coverage"],
                "warning": 4.0,
                "breach": 2.5,
                "direction": "higher_better",
            }
        )

    return {"drivers": drivers, "risks": risks, "catalysts": []}


def extract_phase1_payload(workbook_path, overrides=None):
    overrides = overrides or {}
    evaluator = WorkbookEvaluator(workbook_path)

    summary_ws = evaluator.wb["Summary"]
    summary_title = summary_ws["A1"].value
    workbook_name = Path(workbook_path).stem
    company_name, ticker = parse_title_for_company(summary_title, workbook_name)

    is_periods, is_rows = read_sheet_rows(evaluator, "IS", header_row=3, start_row=4)
    bs_periods, bs_rows = read_sheet_rows(evaluator, "BS", header_row=3, start_row=4)
    cf_periods, cf_rows = read_sheet_rows(evaluator, "CF", header_row=3, start_row=4)
    returns_periods, returns_rows = read_sheet_rows(evaluator, "Returns", header_row=3, start_row=4)
    actuals = actual_periods(is_periods)
    forecasts = forecast_periods(is_periods)

    latest_actual = actuals[-1]
    previous_actual = actuals[-2] if len(actuals) >= 2 else actuals[-1]
    first_forecast = forecasts[0] if forecasts else None
    last_forecast = forecasts[-1] if forecasts else None

    revenue_latest = float(is_rows["Revenue"][latest_actual])
    revenue_prev = float(is_rows["Revenue"][previous_actual]) if previous_actual else None
    ebit_latest = float(is_rows["EBIT"][latest_actual])
    gross_margin_latest = float(is_rows["Gross Margin %"][latest_actual])
    net_income_latest = float(is_rows["Net Income"][latest_actual])
    da_latest = float(is_rows.get("D&A Memo", {}).get(latest_actual) or cf_rows.get("D&A", {}).get(latest_actual) or 0.0)
    cfo_latest = float(cf_rows["CFO"][latest_actual])
    cfi_latest = float(cf_rows["CFI"][latest_actual])
    fcf_latest = cfo_latest + cfi_latest
    total_assets_latest = float(bs_rows["Total Assets"][latest_actual])
    total_liabilities_latest = float(bs_rows["Total Liabilities"][latest_actual])
    current_debt_latest = float(bs_rows["Current Debt"][latest_actual])
    lt_debt_latest = float(bs_rows["LT Debt"][latest_actual])
    cash_latest = float(bs_rows["Cash"][latest_actual])
    roe_latest = float(returns_rows.get("ROE", {}).get(latest_actual) or 0.0)
    roic_latest = float(returns_rows.get("ROIC", {}).get(latest_actual) or 0.0)
    net_debt_latest = float(returns_rows.get("Net Debt", {}).get(latest_actual) or (current_debt_latest + lt_debt_latest - cash_latest))
    ebitda_latest = ebit_latest + da_latest

    revenue_growth_ltm = pct_change(revenue_latest, revenue_prev) if revenue_prev else None
    ebit_margin_ltm = safe_div(ebit_latest, revenue_latest)
    fcf_margin_ltm = safe_div(fcf_latest, revenue_latest)
    net_debt_to_ebitda = safe_div(net_debt_latest, ebitda_latest)
    interest_coverage = safe_div(ebit_latest, abs(float(is_rows["Interest Expense"][latest_actual]))) if is_rows["Interest Expense"][latest_actual] else None

    revenue_cagr_forecast = None
    revenue_growth_next_year = None
    ebit_margin_terminal = None
    fcf_margin_terminal = None
    roic_terminal = None

    if first_forecast and last_forecast:
        revenue_first_forecast = float(is_rows["Revenue"][first_forecast])
        revenue_last_forecast = float(is_rows["Revenue"][last_forecast])
        periods_between = max(1, len(forecasts) - 1)
        if revenue_first_forecast > 0:
            revenue_cagr_forecast = (revenue_last_forecast / revenue_first_forecast) ** (1 / periods_between) - 1
        revenue_growth_next_year = pct_change(revenue_first_forecast, revenue_latest)
        ebit_margin_terminal = safe_div(float(is_rows["EBIT"][last_forecast]), revenue_last_forecast)
        fcf_terminal = float(cf_rows["CFO"][last_forecast]) + float(cf_rows["CFI"][last_forecast])
        fcf_margin_terminal = safe_div(fcf_terminal, revenue_last_forecast)
        roic_terminal = float(returns_rows.get("ROIC", {}).get(last_forecast) or 0.0)

    cross_check_ws = evaluator.wb["Cross_Check"]
    concentration_values = []
    for row in range(1, cross_check_ws.max_row + 1):
        row_text = " ".join(
            str(cross_check_ws.cell(row, col).value or "")
            for col in range(1, cross_check_ws.max_column + 1)
        )
        if "concentration" in row_text.lower():
            concentration_values.extend(extract_percentages(row_text))
    concentration_values = sorted(set(concentration_values), reverse=True)

    company = {
        "name": company_name,
        "ticker": ticker,
        "sector_family": overrides.get("sector_family"),
        "industry": overrides.get("industry"),
        "currency": overrides.get("currency", "USD"),
        "report_date": overrides.get("report_date"),
    }

    model_summary = {
        "latest_actual_period": latest_actual,
        "first_forecast_period": first_forecast,
        "last_forecast_period": last_forecast,
        "revenue_ltm": revenue_latest,
        "revenue_growth_ltm": revenue_growth_ltm,
        "gross_margin_ltm": gross_margin_latest,
        "ebit_margin_ltm": ebit_margin_ltm,
        "fcf_margin_ltm": fcf_margin_ltm,
        "roe_ltm": roe_latest,
        "roic_ltm": roic_latest,
        "net_debt_to_ebitda": net_debt_to_ebitda,
        "interest_coverage": interest_coverage,
        "customer_concentration_top1": concentration_values[0] if concentration_values else None,
        "customer_concentration_top5": sum(concentration_values[:5]) if concentration_values else None,
        "revenue_growth_next_year": revenue_growth_next_year,
        "revenue_cagr_forecast": revenue_cagr_forecast,
        "ebit_margin_terminal": ebit_margin_terminal,
        "fcf_margin_terminal": fcf_margin_terminal,
        "roic_terminal": roic_terminal,
    }

    burn = abs(fcf_latest) if fcf_latest < 0 else None
    cash_runway_months = (cash_latest / burn) * 12 if burn else None
    quality_inputs = {
        "roe_history": series_values(returns_rows, "ROE", actuals),
        "roic_history": series_values(returns_rows, "ROIC", actuals),
        "debt_to_assets": safe_div(total_liabilities_latest, total_assets_latest),
        "net_debt_to_ebitda": net_debt_to_ebitda,
        "interest_coverage": interest_coverage,
        "fcf_to_net_income": safe_div(fcf_latest, net_income_latest),
        "fcf_margin": fcf_margin_ltm,
        "gross_margin": gross_margin_latest,
        "operating_margin": ebit_margin_ltm,
        "moat_types": overrides.get("moat_types", []),
        "cash_runway_months": cash_runway_months,
    }

    monitoring_inputs = build_default_monitoring(model_summary)
    if overrides.get("monitoring_inputs"):
        deep_update(monitoring_inputs, overrides["monitoring_inputs"])

    payload = {
        "company": company,
        "model_summary": model_summary,
        "quality_inputs": quality_inputs,
        "monitoring_inputs": monitoring_inputs,
    }
    if overrides:
        deep_update(payload, overrides)
    return payload


def convert_enterprise_values_to_share_prices(dcf_payload):
    sensitivity = dcf_payload.get("sensitivity")
    if not sensitivity:
        return None, None
    adjustment = get_num(dcf_payload.get("ev_bridge", {}), "enterprise_to_equity_adjustment_total", 0.0)
    diluted_shares = get_num(dcf_payload.get("share_bridge", {}), "diluted_shares")
    if not diluted_shares:
        return None, None
    values = []
    for row in sensitivity.get("matrix", []):
        for value_row in row.get("values", []):
            enterprise_value = get_num(value_row, "enterprise_value")
            if enterprise_value is None:
                continue
            values.append((enterprise_value + adjustment) / diluted_shares)
    if not values:
        return None, None
    return min(values), max(values)


def method_summary_from_dcf(payload):
    mid = get_num(payload, "equity_value_per_share")
    low, high = convert_enterprise_values_to_share_prices(payload)
    if low is None or high is None:
        low = mid * 0.90 if mid is not None else None
        high = mid * 1.10 if mid is not None else None
    return {"name": "DCF", "low": low, "mid": mid, "high": high}


def method_summary_from_comps(payload):
    values = [
        get_num(method, "implied_value_per_share")
        for method in payload.get("methods", [])
        if method.get("implied_value_per_share") is not None
    ]
    values = [value for value in values if value is not None]
    if not values:
        return None
    ordered = sorted(values)
    mid = ordered[len(ordered) // 2] if len(ordered) % 2 == 1 else (ordered[len(ordered) // 2 - 1] + ordered[len(ordered) // 2]) / 2
    return {"name": "Comps", "low": ordered[0], "mid": mid, "high": ordered[-1]}


def reverse_dcf_summary(payload):
    if not payload:
        return {}
    result = dict(payload)
    solved_value = payload.get("solved_value")
    solve_for = payload.get("solve_for")
    mapping = {
        "terminal_growth_rate": "market_implied_growth",
        "terminal_multiple": "market_implied_multiple",
        "discount_rate": "market_implied_discount_rate",
        "cash_flow_multiplier": "market_implied_cash_flow_multiplier",
    }
    if solve_for in mapping and solved_value is not None:
        result[mapping[solve_for]] = solved_value
    return result


def weighted_range(primary, secondary, primary_weight=0.6):
    if primary and secondary:
        secondary_weight = 1.0 - primary_weight
        return {
            "low": primary["low"] * primary_weight + secondary["low"] * secondary_weight,
            "mid": primary["mid"] * primary_weight + secondary["mid"] * secondary_weight,
            "high": primary["high"] * primary_weight + secondary["high"] * secondary_weight,
            "weights": {primary["name"]: primary_weight, secondary["name"]: secondary_weight},
        }
    method = primary or secondary
    if not method:
        return None
    return {
        "low": method["low"],
        "mid": method["mid"],
        "high": method["high"],
        "weights": {method["name"]: 1.0},
    }


def extract_phase2_payload(artifacts, overrides=None, primary_weight=0.6):
    overrides = overrides or {}
    if isinstance(artifacts, (str, Path)):
        artifacts = discover_phase2_files(artifacts)

    if artifacts.get("valuation_summary"):
        valuation = load_json(artifacts["valuation_summary"])
        qc_payload = {}
        if artifacts.get("valuation_qc"):
            qc_payload = load_json(artifacts["valuation_qc"])
        elif artifacts.get("financials_qc"):
            qc_payload = load_json(artifacts["financials_qc"])
        elif artifacts.get("reit_qc"):
            qc_payload = load_json(artifacts["reit_qc"])
        elif artifacts.get("regulated_qc"):
            qc_payload = load_json(artifacts["regulated_qc"])
        elif artifacts.get("asset_nav_qc"):
            qc_payload = load_json(artifacts["asset_nav_qc"])
        elif artifacts.get("biotech_qc"):
            qc_payload = load_json(artifacts["biotech_qc"])
        elif artifacts.get("sotp_qc"):
            qc_payload = load_json(artifacts["sotp_qc"])
        if valuation.get("valuation_qc_passed") is None and qc_payload:
            valuation["valuation_qc_passed"] = bool(qc_payload.get("passed", False))
        valuation.setdefault("artifact_paths", artifacts)
        if overrides:
            deep_update(valuation, overrides)
        return valuation

    dcf = load_json(artifacts["dcf"]) if artifacts.get("dcf") else {}
    comps = load_json(artifacts["comps"]) if artifacts.get("comps") else {}
    reverse_dcf = load_json(artifacts["reverse_dcf"]) if artifacts.get("reverse_dcf") else {}
    football = load_json(artifacts["football_field"]) if artifacts.get("football_field") else {}
    valuation_qc = load_json(artifacts["valuation_qc"]) if artifacts.get("valuation_qc") else {}

    dcf_method = method_summary_from_dcf(dcf) if dcf else None
    comps_method = method_summary_from_comps(comps) if comps else None

    if football.get("methods"):
        by_name = {row.get("name", row.get("label")): row for row in football["methods"]}
        if dcf_method and "DCF" in by_name:
            dcf_method = {
                "name": "DCF",
                "low": get_num(by_name["DCF"], "low", dcf_method["low"]),
                "mid": get_num(by_name["DCF"], "mid", dcf_method["mid"]),
                "high": get_num(by_name["DCF"], "high", dcf_method["high"]),
            }
        if comps_method and "Comps" in by_name:
            comps_method = {
                "name": "Comps",
                "low": get_num(by_name["Comps"], "low", comps_method["low"]),
                "mid": get_num(by_name["Comps"], "mid", comps_method["mid"]),
                "high": get_num(by_name["Comps"], "high", comps_method["high"]),
            }

    combined = weighted_range(dcf_method, comps_method, primary_weight=primary_weight)
    current_price = (
        overrides.get("current_price")
        or football.get("current_price")
        or artifacts.get("current_price")
    )
    valuation = {
        "current_price": current_price,
        "weighted_target_price": combined["mid"] if combined else None,
        "expected_value_per_share": combined["mid"] if combined else None,
        "base_target_price": combined["mid"] if combined else None,
        "bull_target_price": combined["high"] if combined else None,
        "bear_target_price": combined["low"] if combined else None,
        "primary_method": "DCF" if dcf_method else ("Comps" if comps_method else None),
        "secondary_method": "Comps" if dcf_method and comps_method else None,
        "method_weights": combined["weights"] if combined else {},
        "valuation_qc_passed": bool(valuation_qc.get("passed", False)) if valuation_qc else None,
        "reverse_dcf": reverse_dcf_summary(reverse_dcf),
        "scenarios": (
            [
                {"name": "bear", "value_per_share": combined["low"], "probability": 0.25},
                {"name": "base", "value_per_share": combined["mid"], "probability": 0.50},
                {"name": "bull", "value_per_share": combined["high"], "probability": 0.25},
            ]
            if combined
            else []
        ),
        "methods": [method for method in [dcf_method, comps_method] if method],
        "artifact_paths": artifacts,
    }
    if overrides:
        deep_update(valuation, overrides)
    return valuation
